"""
set_excel.py

只负责创建 Excel 的 config sheet 和费率表 sheet
从 get_content_vl 模块获取配置数据
从 1.py 生成的 txt 文件读取费率表数据
"""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter

# 从 set_config 模块导入配置数据（动态获取）
try:
    import set_config
except ImportError:
    set_config = None


# ── 费率表 TXT 解析 ──────────────────────────────────────────────────────────

def parse_rate_txt(txt_path: Path) -> dict | None:
    """
    解析 1.py 生成的费率表 txt 文件，支持多 section（多被保人分段）。

    返回结构:
        {
            "sections": [
                {
                    "headers": [("保险期间", [...]), ("性别", [...]), ...],
                    "data":    [(age_str, [val, ...]), ...],
                },
                ...
            ]
        }
    如果文件无法解析返回 None。
    """
    lines = [line.strip() for line in txt_path.read_text(encoding="utf-8").splitlines() if line.strip()]
    if not lines:
        return None

    sections: list[dict] = []
    current_headers: list[tuple[str, list[str]]] = []
    current_data: list[tuple[str, list[str]]] = []
    in_data = False

    for line in lines:
        parts = line.split(",")
        first = parts[0].strip()

        if first.isdigit():
            in_data = True
            vals = [v.strip() for v in parts[1:]]
            current_data.append((first, vals))
        else:
            # 非数字行：如果当前已有数据，说明是新 section 开始
            if in_data and current_data:
                sections.append({"headers": current_headers, "data": current_data})
                current_headers = []
                current_data = []
                in_data = False
            if len(parts) >= 2:
                current_headers.append((first, [v.strip() for v in parts[1:]]))

    # 收集最后一个 section
    if current_headers or current_data:
        sections.append({"headers": current_headers, "data": current_data})

    return {"sections": sections} if sections else None


def _find_header(headers: list[tuple[str, list[str]]], key: str) -> list[str] | None:
    """在 headers 列表中查找指定 key，返回对应的值列表；未找到返回 None。"""
    for k, v in headers:
        if k == key:
            return v
    return None


def _normalize_payment(period: str) -> str:
    """将交费期间格式化为 'x年'（去掉末尾的 交/缴/期），'一次交清' 保持不变。"""
    import re
    if re.match(r"^一次[性]?[交缴][纳清]$", period):
        return "一次交清"
    return re.sub(r"(年)(?:交|缴|期)$", r"\1", period)


_SUB_GROUP_DISPLAY = {"一人": "被保人为一人", "两人": "被保人为两人"}


def _format_sub_group(val: str) -> str:
    """将 '一人'/'两人' 转换为 '被保人为一人'/'被保人为两人'，其他值保持不变。"""
    return _SUB_GROUP_DISPLAY.get(val, val)


# ── 费率表 sheet 填充 ────────────────────────────────────────────────────────

def fill_rate_sheet(ws, txt_path: Path) -> bool:
    """
    读取 txt 文件并填充费率表 sheet，支持多 section 数据。

    多 section 策略：
        - 如果各 section 年龄集合相同 → 横向合并（两人放一人右边）
        - 否则纵向堆叠

    每个 section 的 header 行顺序：
        保险期间（无则"终身"）→ 交费期间 → 投保计划/可选责任（多被保人）→ 性别
    """
    parsed = parse_rate_txt(txt_path)
    if parsed is None or not parsed["sections"]:
        print(f"⚠️ 无法解析费率表数据: {txt_path.name}")
        return False

    sections = parsed["sections"]

    # ── 判断是否可以横向合并 ──────────────────────────────────────────────────
    can_merge = len(sections) > 1 and all(
        [age for age, _ in sec["data"]] == [age for age, _ in sections[0]["data"]]
        and len(sec["data"]) == len(sections[0]["data"])
        for sec in sections[1:]
    )

    # ── 样式 ──────────────────────────────────────────────────────────────────
    header_fill = PatternFill(start_color="8FCC4F", end_color="8FCC4F", fill_type="solid")
    age_fill    = PatternFill(start_color="F8CC44", end_color="F8CC44", fill_type="solid")
    data_fill   = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin"),
    )
    center_align = Alignment(horizontal="center", vertical="center")

    def write_row(row_idx: int, values: list, fill: PatternFill = header_fill) -> None:
        for col_idx, val in enumerate(values, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = fill
            cell.border    = thin_border
            cell.alignment = center_align

    def write_data_row(row_idx: int, age: str, values: list) -> None:
        """数据行：年龄列用 age_fill，数值列用 data_fill。"""
        cell = ws.cell(row=row_idx, column=1, value=age)
        cell.fill      = age_fill
        cell.border    = thin_border
        cell.alignment = center_align
        for col_idx, val in enumerate(values, start=2):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill      = data_fill
            cell.border    = thin_border
            cell.alignment = center_align

    data_start_row = 1

    if can_merge:
        # ── 横向合并：所有 section 拼到同一行 ────────────────────────────────
        num_cols = sum(
            len(_find_header(sec["headers"], "投保年龄|交费期间") or [])
            for sec in sections
        )

        def merged_header(key: str, label: str, default_vals: list[str] | None = None) -> list[str] | None:
            """将所有 section 的同名 header 值拼接在一起。"""
            combined: list[str] = []
            for sec in sections:
                vals = _find_header(sec["headers"], key)
                if vals is not None:
                    combined.extend(vals)
                elif default_vals is not None:
                    n = len(_find_header(sec["headers"], "投保年龄|交费期间") or [])
                    combined.extend(default_vals[:n] if len(default_vals) >= n else default_vals * (n // max(len(default_vals), 1) + 1))
            return combined or None

        # 保险期间
        ip_vals = []
        for sec in sections:
            v = _find_header(sec["headers"], "保险期间")
            n = len(_find_header(sec["headers"], "投保年龄|交费期间") or [])
            ip_vals.extend(v if v else ["终身"] * n)
        current_row = 1
        write_row(current_row, ["保险期间"] + ip_vals)
        current_row += 1

        # 交费期间
        pp_vals = []
        for sec in sections:
            v = _find_header(sec["headers"], "投保年龄|交费期间")
            if v:
                pp_vals.extend(_normalize_payment(x) for x in v)
        if pp_vals:
            write_row(current_row, ["交费期间"] + pp_vals)
            current_row += 1

        # 投保计划 / 可选责任（兼容"多被保人"、"投保计划"和"可选责任"三种表头）
        ni_vals = []
        ni_label = "投保计划"
        for sec in sections:
            v = _find_header(sec["headers"], "多被保人")
            if v is None:
                v = _find_header(sec["headers"], "投保计划")
                if v is None:
                    v = _find_header(sec["headers"], "可选责任")
                    if v is not None:
                        ni_label = "可选责任"
                else:
                    ni_label = "投保计划"
            else:
                ni_label = "投保计划"
            if v is not None:
                ni_vals.extend(v)
        if ni_vals:
            write_row(current_row, [ni_label] + [_format_sub_group(v) for v in ni_vals])
            current_row += 1

        # 性别
        g_vals = []
        for sec in sections:
            v = _find_header(sec["headers"], "性别")
            if v is not None:
                g_vals.extend(v)
        if g_vals:
            write_row(current_row, ["性别"] + g_vals)
            current_row += 1

        # 数据行：按年龄横向拼接
        data_start_row = current_row
        base_data = sections[0]["data"]
        for i, (age, _) in enumerate(base_data):
            merged_vals: list[str] = []
            for sec in sections:
                if i < len(sec["data"]):
                    merged_vals.extend(sec["data"][i][1])
            write_data_row(current_row, age, merged_vals)
            current_row += 1

        total_data_rows = len(base_data)

    else:
        # ── 纵向堆叠（单 section 或年龄不同） ────────────────────────────────
        current_row = 1
        total_data_rows = 0
        num_cols = 0

        for section in sections:
            headers = section["headers"]
            data    = section["data"]

            insurance_periods = _find_header(headers, "保险期间")
            payment_periods   = _find_header(headers, "投保年龄|交费期间")
            num_insured       = _find_header(headers, "多被保人")
            plan_vals         = _find_header(headers, "投保计划")
            optional_vals     = _find_header(headers, "可选责任")
            genders           = _find_header(headers, "性别")

            num_cols = max(num_cols, len(payment_periods) if payment_periods else 0)
            if num_cols == 0:
                continue

            # 保险期间（无则"终身"）
            if insurance_periods:
                write_row(current_row, ["保险期间"] + list(insurance_periods))
            else:
                write_row(current_row, ["保险期间"] + ["终身"] * num_cols)
            current_row += 1

            # 交费期间
            if payment_periods:
                write_row(current_row, ["交费期间"] + [_normalize_payment(x) for x in payment_periods])
                current_row += 1

            # 投保计划 / 性别
            handled = {"保险期间", "投保年龄|交费期间"}

            if num_insured is not None:
                write_row(current_row, ["投保计划"] + [_format_sub_group(v) for v in num_insured])
                current_row += 1
                handled.add("多被保人")
            elif plan_vals is not None:
                write_row(current_row, ["投保计划"] + list(plan_vals))
                current_row += 1
                handled.add("投保计划")
            elif optional_vals is not None:
                write_row(current_row, ["可选责任"] + list(optional_vals))
                current_row += 1
                handled.add("可选责任")

            if genders is not None:
                write_row(current_row, ["性别"] + list(genders))
                current_row += 1
                handled.add("性别")

            # 其余未知 header（未来扩展兼容）
            for key, vals in headers:
                if key not in handled:
                    write_row(current_row, [key] + list(vals))
                    current_row += 1

            # 数据行
            if data_start_row == 1:
                data_start_row = current_row
            for age, values in data:
                write_data_row(current_row, age, list(values))
                current_row += 1
                total_data_rows += 1

    # 冻结表头行 + 年龄列
    ws.freeze_panes = f"B{data_start_row}"

    # 自动列宽
    ws.column_dimensions["A"].width = 14
    for i in range(2, num_cols + 2):
        ws.column_dimensions[get_column_letter(i)].width = 13

    print(f"✅ 费率表已填充: {txt_path.name} → {total_data_rows} 行数据，{num_cols} 列")
    return True


# ── 主入口 ───────────────────────────────────────────────────────────────────

def create_config_sheet(output_path=None, txt_path=None):
    """
    创建包含 config sheet 和费率表 sheet 的 Excel 文件。

    Args:
        output_path: 输出 Excel 文件路径（None 时使用 product_name 作为文件名）
        txt_path:    1.py 生成的费率表 txt 文件路径；
                     None 时自动扫描脚本所在目录的 txt 文件。
    """
    # 动态获取最新的 config_data
    if set_config is not None and hasattr(set_config, 'config_data'):
        config_info = set_config.config_data
        print('✅ 从 set_config 获取配置数据')
    else:
        config_info = {}
        print('⚠️ 未找到配置数据，使用空值')

    # 如果没有提供 output_path，使用 product_name 作为文件名
    if output_path is None:
        product_name = config_info.get('product_name', '') if config_info else ''
        if product_name:
            safe_name = (product_name
                         .replace('/', '_').replace('\\', '_')
                         .replace(':', '_').replace('*', '_')
                         .replace('?', '_').replace('"', '_')
                         .replace('<', '_').replace('>', '_')
                         .replace('|', '_'))
            output_path = f"{safe_name}.xlsx"
            print(f'📝 使用产品名称作为文件名: {output_path}')
        else:
            output_path = 'output.xlsx'
            print('⚠️ 未找到产品名称，使用默认文件名: output.xlsx')

    output_path = Path(output_path)

    # ── 创建 workbook ─────────────────────────────────────────────────────────
    wb = Workbook()
    ws_config = wb.active
    ws_config.title = 'config'

    # 创建费率表 sheet
    ws_rate = wb.create_sheet(title='费率表')

    # ── 写入 config sheet ─────────────────────────────────────────────────────
    def safe_get(key, default=''):
        return config_info.get(key, default) if config_info else default

    # 根据 txt 表头是否含有"多被保人"、"投保计划"或"可选责任"动态设置 start_row 和 config_params
    has_sub_dimension = False
    sub_dimension_name = ''
    if txt_path is not None:
        try:
            txt_content = Path(txt_path).read_text(encoding='utf-8')
            if '多被保人' in txt_content or '投保计划' in txt_content:
                has_sub_dimension = True
                sub_dimension_name = '投保计划'
            elif '可选责任' in txt_content:
                has_sub_dimension = True
                sub_dimension_name = '可选责任'
        except Exception:
            pass

    start_row = 5 if has_sub_dimension else 4
    config_params = '年龄,性别,保险期间,交费期间' + (',' + sub_dimension_name if has_sub_dimension else '')

    config_rows = [
        ['产品名称', safe_get('product_name')],
        ['产品编码', safe_get('product_code')],
        ['计算单位', safe_get('calculation_unit')],
        ['开始行', start_row],
        ['开始列', 2],
        ['年龄', '纵向'],
        ['需要参数', config_params],
        ['交费方式', '年交'],
        ['月缴基数', safe_get('月缴基数')],
        ['季缴基数', safe_get('季缴基数')],
        ['半年缴基数', safe_get('半年缴基数')],
        ['保费参数', safe_get('premium_type')],
    ]

    green_fill  = PatternFill(start_color="8FCC4F", end_color="8FCC4F", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'),
    )

    for row_idx, row_data in enumerate(config_rows, start=1):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_config.cell(row=row_idx, column=col_idx, value=value)
            cell.fill      = green_fill
            cell.border    = thin_border
            cell.alignment = Alignment(horizontal='left', vertical='center')

    ws_config.column_dimensions['A'].width = 20
    ws_config.column_dimensions['B'].width = 50

    # ── 填充费率表 sheet ──────────────────────────────────────────────────────
    if txt_path is not None:
        fill_rate_sheet(ws_rate, Path(txt_path))
    else:
        # 自动扫描脚本所在目录
        project_dir = Path(__file__).resolve().parent
        candidates = sorted(project_dir.glob("*.txt"))
        filled = False
        for txt_file in candidates:
            if fill_rate_sheet(ws_rate, txt_file):
                filled = True
                break
        if not filled:
            print("⚠️ 未找到可用的费率表 txt 文件，费率表 sheet 为空")

    # ── 保存 ──────────────────────────────────────────────────────────────────
    wb.save(output_path)
    print(f'✅ 已保存 Excel: {output_path}')
